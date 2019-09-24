#coding:utf-8
import requests
import json
import string
import re
from random import choice
import xlrd
import xlwt
import openpyxl
import unittest
import HTMLTestRunner
import time

class UrlTest(unittest.TestCase):
    testcase = xlrd.open_workbook("E:\JackyYang\jacky.xlsx")  # 打开某个地址的表(修改点)
    sheet = testcase.sheet_by_index(0)  # 打开第一张索引的表

    "该函数目的在于获取EXCEL文件中的URL地址"
    def geturl(self):
        url = self.sheet.cell_value(rowx=1, colx=0)  #获取URL链接地址
        print(url)
        return url

    "将参数拼装后使用request库进行请求"
    def test_advice(self):
        url = self.geturl() #首先获取到需要测试的URL地址

        for i in range(1,self.sheet.nrows):  #每一行逐行进行获取数据
            sheet_cow = self.sheet.row_values(rowx=i) #获取每行的数据
            sheet_parameter =sheet_cow[1:-2]    #第一列与倒数两列的数据是不要的，所以要分割一下
            # print(sheet_parameter)
            content = sheet_parameter[0]    #根据对应的参数类型 进行赋值（修改点）
            image = sheet_parameter[1]      #根据对应的参数类型 进行赋值（修改点）
            qq = int(sheet_parameter[2])    #根据对应的参数类型 进行赋值（修改点）
            data = {'content': content,'image':image,'qq':qq}  #ulr request 执行
            response = requests.post(url, data=data)
            data_test = response.text
            r = json.loads(response.text)
            getresponsetime = response.elapsed.microseconds #返回的服务器响应时间，返回的为微秒，转换为秒
            datalist = re.findall(r'\"ret\":.*?(?=,)', data_test)
            print(f'获取到的响应状态为：{datalist[0]}')
            print(f'服务器响应时间为:{getresponsetime/1000000}')
            self.assertEqual(datalist[0], '"ret":0')
            wb = openpyxl.load_workbook('E:\JackyYang\jacky.xlsx')  #返回的数据写入
            ws = wb.active
            ws.cell(row=i + 1, column=((self.sheet.ncols)-1),value= str(r)) #写入返回数据，数据通过json.loads转为python格式，然后用字符串写入保存
            ws.cell(row=i + 1, column=self.sheet.ncols, value=str(getresponsetime/1000000))  #写入返回的服务器响应时间，返回的为微秒，转换为秒
            wb.save("E:\JackyYang\jacky.xlsx")


"测试组件，可按要求进行修改名称即可"
def suite():
    adviceTest = unittest.makeSuite(UrlTest, "test")
    return adviceTest

if __name__ == '__main__':
    now = time.strftime("%Y-%m-%d %H-%M-%S")
    filename = 'E:/JackyYang/TestReport/'+ now + 'XXXX接口测试报告.html'  #名称可随意修改，包括路径
    report = open(filename, "wb")
    runner = HTMLTestRunner.HTMLTestRunner(stream=report, title="测试报告", description="测试报告详情")
    runner.run(suite())
    report.close()